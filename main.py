from dataclasses import dataclass
from math import sqrt
from classes import student_data, Student
from random import randint
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import imageio.v2 as iio
import openpyxl as xl

q1 = str(input("\n Have you uploaded your csv data file with name 'original_data.csv' to img folder? y/n "))

if q1 == "n":
    print("\n Please complete this step before running this program.")
    exit()

q2 = str(input("\n Have you processed your data by running 'scripts/data_process.py'? y/n "))

if q2 == "n":
    print("\n Please complete this step before running this program.")
    exit()


print("\n Input integer 2-6 to form the integer number of clusters to analyze your data.\n")

k = int(input(" Number of clusters > "))

@dataclass #Decorators
class Point:
    sub_id: int
    sub1: float
    sub2: float
    sub3: float 
    center_index = -1

@dataclass
class Cluster:
    sub1: int
    sub2: int
    sub3: float 
    previous_sub1 = -1.0
    previous_sub2 = -1.0
    previous_sub3 = -1.0

    def has_moved(self, e = 1e-4) -> bool:
        d = sqrt(
            (self.previous_sub1 - self.sub1)**2 
            + (self.previous_sub2 - self.sub2)**2
            + (self.previous_sub3 - self.sub3)**2
        )
        self.previous_sub1, self.previous_sub2, self.previous_sub3 = self.sub1, self.sub2, self.sub3
        return d > e


def random_centroids(k, data) -> list[Cluster]:
    """Initate the centroids by using random coordinates within the dataset"""

    centroids = []
    for _ in range (k):
        
        centroids.append(Cluster(data[randint(0,len(data) - 1)].sub1, data[randint(0,len(data) - 1)].sub2, data[randint(0,len(data) - 1)].sub3))
        
    return centroids

def assign_cluster(centroids, data) -> list[list[Cluster],int]:
    """Find the nearest cluster of a single point, assign point to the cluster, repeat for all"""

    points_with_cluster = []
    for point in data:
        distance = []
        count = 0 #index for centroids
        for center in centroids:
            distance.append([Student.distance(center, point), count])
            count += 1
        distance.sort(key = lambda x: x[0]) 
        points_with_cluster.append([point,distance[0][1]])

    return points_with_cluster

def move_centroids(k, centroids, points_with_cluster) -> None:
    """Find the average x&y of the points, make that position the new centroid"""

    for i in range(k):
        total_x = 0
        total_y = 0
        total_z = 0
        count = 0
        for point in points_with_cluster:
            if point[1] == i:
                total_x += point[0].sub1
                total_y += point[0].sub2
                total_z += point[0].sub3
                count += 1
        if count == 0:
            main()
        centroids[i].sub1 = total_x/count # Cluster() is used to create a new cluster point. Not covering the initial centroid point
        centroids[i].sub2 = total_y/count
        centroids[i].sub3 = total_z/count

def plotting(k, centroids, points_with_cluster) -> None:
    """Plot the sepal width and petal width """

    fig = plt.figure(figsize = plt.figaspect(1) )
    ax = plt.axes(projection='3d')
    x = [centroid.sub1 for centroid in centroids]
    y = [centroid.sub2 for centroid in centroids]
    z = [centroid.sub3 for centroid in centroids]
    ax.scatter(x, y, z, marker='x', color='black')

    colors = ['red','green','blue','brown','purple','orange']

    for point in points_with_cluster:

        x1 = [point[0].sub1]
        y1 = [point[0].sub2]
        z1 = [point[0].sub3]

        for i in range(k):
            if i == point[1]:
                chosen_color = colors[i]

        ax.scatter(x1, y1, z1, marker = "o", color = chosen_color, alpha=0.4)

        clusters = []
        for i in range (k):
            clusters.append(mpatches.Patch(color = colors[i], label = f'Cluster {i+1}'))
        plt.legend(handles=clusters)

def save_excel(k, points_with_cluster):
    """Save excel files of the data points included in peach cluster"""

    for i in range(k):
        wb = xl.Workbook()
        ws = wb.active

        ws["A1"] = "id"
        ws["B1"] = "lang_z"
        ws["C1"] = "stem_z"
        ws["D1"] = "hum_z"
        
        id_ = []
        lang_z = []
        stem_z = []
        hum_z = []

        temp = 0

        for point in points_with_cluster:
            if point[1] == i:
                id_.append(point[0].sub_id)
                lang_z.append(point[0].sub1)
                stem_z.append(point[0].sub2)
                hum_z.append(point[0].sub3)
                temp += 1

        for count in range(temp):
            ws[f'A{count+2}'] = id_[count]
            ws[f'B{count+2}'] = lang_z[count]
            ws[f'C{count+2}'] = stem_z[count]
            ws[f'D{count+2}'] = hum_z[count]


        wb.save(f"docs/Cluster{i+1}.xlsx")

def save_gif(img_paths: list[str], path: str) -> None:
    """Create a gif animation of the pictures generated"""

    imgs = [iio.imread(img_path) for img_path in img_paths]
    iio.mimsave(path, imgs, duration = 0.2)
    print("\n Animation file saved at img folder as 'animate.gif.'")

def k_means(k: int, data: list[Point]):
    """Performs k-means on the data, plots the points and clusters after each iteration"""
    
    counter = 0
    centroids = random_centroids(k, data)
    paths = []

    # continously move the centroids until the distance they are being moved by is negligible
    while any([c.has_moved() == True for c in centroids]): # Any => if one element in a list is true, the function returns true

        points_with_cluster = assign_cluster(centroids, data)

        # Plotting initial graph and each iteration
        plotting(k, centroids, points_with_cluster)

        path = f'img/G{counter+1}.png'
        plt.savefig(path)
        paths.append(path)
        counter += 1
        


        move_centroids(k, centroids, points_with_cluster)
    print("---------------------------------")
    print(f"\n Graphs 1 to {counter} saved at img folder as .png file. ")
    save_gif(paths, 'img/animate.gif')
    save_excel(k, points_with_cluster)
    print(f"\n Excel files clusters 1 to {k} saved at doc as .xlsx file. \n")
    print(" You may now access these files in their corresponding folder. \n")
    exit()


def main():
    """Runs the Program"""

    #inputs

    points = [
        Point(
            student.sub_id,
            student.sub1,
            student.sub2,
            student.sub3
        )
        for student in student_data
    ]
    k_means(k, points)
    

main()
